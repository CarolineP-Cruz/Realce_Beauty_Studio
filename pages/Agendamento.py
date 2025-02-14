# Contents of ~/my_app/pages/page_2.py
import streamlit as st
import pandas as pd

st.markdown("# Agendamento 💇‍♀️ 💅")
st.sidebar.image('logo2.png', use_container_width=True)

import pandas as pd

arquivo_excel = "marcações.xlsx"
arquivo_excel2 = "serviços.xlsx"

# Ler o arquivo Excel
dfs = pd.read_excel(arquivo_excel, sheet_name=None, engine="openpyxl")

# Dicionário para armazenar a saída Markdown de cada planilha
markdown_output = {}

for nome_planilha, df in dfs.items():
    # Converter cada DataFrame para Markdown e armazenar no dicionário
    markdown_output[nome_planilha] = df.to_markdown(index=False)

# Criar saída formatada
marcação = "# Relatório das Planilhas\n\n"

for nome, md in markdown_output.items():
    marcação += f"## Planilha: {nome}\n\n"
    marcação += md + "\n\n"
    marcação += "---\n\n"

import pandas as pd

df =  pd.read_excel(arquivo_excel2)

#Converte o DataFrame para uma tabela Markdown (sem o índice)
informação = df.to_markdown(index=False)

from openai import OpenAI
client = OpenAI(api_key= st.secrets["OPENAI_API_KEY"])

sistema = f"""
---

**Contexto:**
Você é uma assistente virtual de atendimento ao cliente para um salão de beleza. Sua função é gerenciar agendamentos de serviços, remarcações, cancelamentos, consultas de preços e responder dúvidas frequentes. Sua atuação deve seguir as orientações abaixo, garantindo um atendimento claro, organizado e focado exclusivamente em agendamentos.

---

### 1. **Tabela de Preços:**

{informação}

---

### 2. Informações Adicionais

- **Horário de Atendimento:**
  - **Manhã:** 08:00 às 12:00
  - **Tarde:** 13:30 às 17:00

- **Localização:**
  Rua dos Tupinambás, nº 580, Bairro Pontal, Ilhéus, Bahia

- **Formas de Pagamento:**
  Cartão, Pix ou Dinheiro
  - Para serviços acima de R$100,00, é possível parcelar em duas vezes.

- **Profissionais Responsáveis:**
  - **Cabelereiro:** Beatriz
  - **Manicure:** Camila
  - **Designer de Sobrancelha:** Daniela
  - **Maquiador:** Elisa


### 3. Verificação de Disponibilidade

- **Planilhas de Agendamentos:**
  Cada profissional possui uma planilha com os agendamentos realizados. Essas planilhas indicam os dias e horários já ocupados.

  As planilhas estam organizadas abaixo:

{marcação}

- **Regras para Agendamento:**
  - Os horários e dias **não listados** na planilha de um profissional estão DISPONÍVEIS.
  - Os agendamentos são permitidos apenas nos meses de **fevereiro, março, abril e maio**.
  - Verifique se o período disponível é suficiente para a duração do serviço (evite sobreposição de horários).
  - O horário de início do novo agendamento não pode ser antes do horário de fim de um agendamento já existente.
  - O horário de fim do novo agendamento não deve se sobrepor ao horário de início de nenhum agendamento existente.
  - Quando o clinte informar um horário, some o horário à duração do serviço e verifique se tem sobreposição de horário da planilha.
  - Verifique se a marcação não ultrapassa o Horário de Atendimento

- **Observação:**
  Cada profissional é considerado individualmente. Assim, é possível que, por exemplo, o cabelereiro esteja ocupado em um horário 
  enquanto a manicure esteja disponível simultaneamente.

---

### 4. Atendimento e Confirmação

- **Processo de Agendamento:**
  1. **Cumprimente** o cliente.
  2. **Identifique** o serviço desejado e informe o preço e a duração com base na tabela.
  3. **Pergunte** a data, o horário preferido e o nome do cliente (se ainda não informado).
  4. **Verifique** a disponibilidade do profissional responsável:
     - Consulte a planilha do profissional para o dia solicitado.
     - Garanta que o intervalo disponível é suficiente para a duração do serviço, sem conflitos com outros agendamentos ou com o horário de descanso.
  5. **Confirme** o agendamento e, se tudo estiver correto, gere um número de protocolo de 6 dígitos, diferente dos outros números de protocolo.
  6. Informe o profissional, data, serviço, horário, nome e protocolo da seguinte maneira nessa ordem:
      **Informações do agendamento**
      Profissional : Cabelereiro (Deve ser a profissão e não o nome do profissional)
      Data: 15/02/2025
      Serviço: Corte feminino
      Horário de Início: 10:30
      Horário de Fim: 10:30
      Nome: Ana
      protocolo: 593684
  7. Informe que o cancelamento ou remarcação só podera ser feita com o número do protocolo.

- **Importante:**
  Nunca forneça nomes de clientes, detalhes de agendamentos ou números de protocolo de outros atendimentos.

---

### 5. Remarcação e Cancelamento

- **Para remarcar ou cancelar:**
  - Solicite que o cliente informe o **número de protocolo** gerado no agendamento.
  - Informe que somente com esse protocolo o serviço pode ser alterado ou cancelado.

---

### 6. Regras para a Conversa

- **Foco:**
  A conversa deve tratar exclusivamente de agendamentos, remarcações, cancelamentos e atendimento ao cliente.

- **Desvios:**
  Se o assunto se desviar do foco principal, repreenda o cliente e redirecione a conversa para os tópicos relacionados aos agendamentos.

- **Sigilo:**
  Não revele informações de outros clientes, agendamentos ja marcados ou protocolos alheios.

---

"""
from IPython.display import Markdown

import re
from openpyxl import load_workbook

# Função para extrair os dados do agendamento da mensagem
def extrair_dados(mensagem):
    # Remover espaços extras e quebras de linha antes de tentar capturar as informações
    mensagem = mensagem.replace('\n', ' ').strip()
    
    # Expressão regular ajustada para capturar as informações do agendamento
    padrao = r"""
    -?\s*Profissional\s*:\s*(?P<profissional>[^:]+?)\s+  # Profissional
    -?\s*Data\s*:\s*(?P<data>\d{2}/\d{2}/\d{4})\s+       # Data
    -?\s*Serviço\s*:\s*(?P<servico>[^:]+?)\s+            # Serviço
    -?\s*Horário\ de\ Início\s*:\s*(?P<inicio>\d{2}:\d{2})\s+  # Horário de início
    -?\s*Horário\ de\ Fim\s*:\s*(?P<fim>\d{2}:\d{2})\s+  # Horário de fim
    -?\s*Nome\s*:\s*(?P<nome>[^:]+?)\s+                 # Nome
    -?\s*Protocolo\s*:\s*(?P<protocolo>\d+)             # Protocolo
    """
    
    match = re.search(padrao, mensagem, re.VERBOSE)


    if match:
        return match.groupdict()
    else:
        return None

# Função para atualizar a planilha com os dados extraídos
def atualizar_planilha(agendamento, file_path="arquivo_excel.xlsx"):
    wb = load_workbook(file_path)
    planilha_profissional = wb[agendamento['profissional']] 

    ultima_linha = planilha_profissional.max_row + 1
    planilha_profissional.cell(row=ultima_linha, column=1, value=agendamento['data'])
    planilha_profissional.cell(row=ultima_linha, column=2, value=agendamento['profissional'])
    planilha_profissional.cell(row=ultima_linha, column=3, value=agendamento['servico'])
    planilha_profissional.cell(row=ultima_linha, column=4, value=agendamento['inicio'])
    planilha_profissional.cell(row=ultima_linha, column=5, value=agendamento['fim'])
    planilha_profissional.cell(row=ultima_linha, column=6, value=agendamento['nome'])
    planilha_profissional.cell(row=ultima_linha, column=7, value=agendamento['protocolo'])

    # Salvar a planilha após atualizar
    wb.save(file_path)
    st.write("Planilha atualizada com sucesso!")

    # Função para acessar a última resposta do chatbot
def ultima_resposta(conversa):
    # A última entrada do histórico de conversa é a resposta do sistema (chatbot)
    resposta = conversa[-1]['content'] if conversa[-1]['role'] == 'system' else None
    return resposta

# Função para verificar a última resposta e atualizar a planilha se necessário
def verificar_e_atualizar(conversa, file_path=arquivo_excel):

    # Acessar a última resposta do chatbot
    resposta_do_chatbot = ultima_resposta(conversa)

    if resposta_do_chatbot and "Informações do agendamento" in resposta_do_chatbot: 
        # Extrair os dados da resposta
        agendamento = extrair_dados(resposta_do_chatbot)

        if agendamento:
            atualizar_planilha(agendamento, file_path)


#############################################################3

# Define o comportamento do assistente e inicializa a conversa
modelo = "gpt-4o-mini"

# Avatares para o usuário e assistente
avatar_user = "💇"  # Substitua pelo caminho correto
avatar_assistent = "💇‍♀️"  # Substitua pelo caminho correto

# Inicializa o histórico da conversa
if "conversa" not in st.session_state:
    st.session_state.conversa = [{'role': 'system', 'content': sistema}]

# Exibe o histórico de mensagens com os avatares correspondentes
for msg in st.session_state.conversa[1:]:
    avatar = avatar_user if msg["role"] == "user" else avatar_assistent
    st.chat_message(msg["role"], avatar=avatar).write(msg["content"])

# Captura a entrada do usuário
prompt = st.chat_input("Digite sua mensagem:")

if prompt:
    # Adiciona a mensagem do usuário ao histórico
    st.session_state.conversa.append({'role': 'user', 'content': prompt})
    st.chat_message("user", avatar=avatar_user).write(prompt)

    # Faz uma requisição à API OpenAI para gerar a resposta do assistente
    with st.chat_message("assistant", avatar=avatar_assistent):
        stream = client.chat.completions.create(
            model=modelo,
            messages=st.session_state.conversa,
            stream=True
        )

        # Exibe a resposta em tempo real
        response = st.write_stream(stream)

    # Adiciona a resposta ao histórico
    st.session_state.conversa.append({'role': 'system', 'content': response})
    
    # Verifica e atualiza a planilha se necessário
    verificar_e_atualizar(st.session_state.conversa, file_path=arquivo_excel)

